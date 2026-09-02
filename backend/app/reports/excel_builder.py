"""Excel workbook report builder using openpyxl."""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.reports.base import BaseReportBuilder


class ExcelReportBuilder(BaseReportBuilder):
    """Builds a multi-tab Excel workbook (.xlsx) for project reporting."""

    def build(self, data: dict[str, Any]) -> bytes:
        wb = openpyxl.Workbook()
        
        # Style definitions
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        sub_font = Font(name="Calibri", size=10, italic=True, color="475569")
        bold_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        bg_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary["A1"] = f"Project Progress Report: {self.project_name}"
        ws_summary["A1"].font = title_font
        ws_summary["A2"] = f"Generated at: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        ws_summary["A2"].font = sub_font

        # Parameters
        ws_summary["A4"] = "Report Parameters"
        ws_summary["A4"].font = bold_font
        ws_summary["A5"] = "Discipline"
        ws_summary["B5"] = str(self.parameters.get("discipline") or "All")
        ws_summary["A6"] = "Status"
        ws_summary["B6"] = str(self.parameters.get("status") or "All")

        # Metrics
        ws_summary["A8"] = "Metric"
        ws_summary["B8"] = "Value"
        for col in ["A8", "B8"]:
            ws_summary[col].fill = header_fill
            ws_summary[col].font = header_font

        summary = data.get("summary", {})
        metric_rows = [
            ("Total Activities", summary.get("total_activities", 0)),
            ("Completed Activities", summary.get("completed_activities", 0)),
            ("In Progress Activities", summary.get("in_progress_activities", 0)),
            ("Delayed Activities", summary.get("delayed_activities", 0)),
            ("Progress Percentage", f"{summary.get('progress_pct', 0.0):.1f}%"),
            ("High Risk Count", summary.get("high_risk_count", 0)),
        ]
        for idx, (label, val) in enumerate(metric_rows, start=9):
            ws_summary[f"A{idx}"] = label
            ws_summary[f"B{idx}"] = val
            ws_summary[f"A{idx}"].border = thin_border
            ws_summary[f"B{idx}"].border = thin_border
            if idx % 2 == 1:
                ws_summary[f"A{idx}"].fill = bg_alt
                ws_summary[f"B{idx}"].fill = bg_alt

        # Sheet 2: Activities
        ws_act = wb.create_sheet(title="Activity Details")
        ws_act.views.sheetView[0].showGridLines = True

        headers = ["Activity Code", "Activity Name", "Discipline", "WBS Path", "Status", "Progress %", "Risk Band"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_act.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        activities = data.get("activities", [])
        for row_idx, act in enumerate(activities, start=2):
            ws_act.cell(row=row_idx, column=1, value=str(act.get("code") or ""))
            ws_act.cell(row=row_idx, column=2, value=str(act.get("name") or ""))
            ws_act.cell(row=row_idx, column=3, value=str(act.get("discipline") or ""))
            ws_act.cell(row=row_idx, column=4, value=str(act.get("wbs_path") or ""))
            ws_act.cell(row=row_idx, column=5, value=str(act.get("status") or "NOT_STARTED"))
            ws_act.cell(row=row_idx, column=6, value=float(act.get("progress_pct", 0.0)))
            ws_act.cell(row=row_idx, column=7, value=str(act.get("risk_band") or "LOW"))

            for c in range(1, 8):
                ws_act.cell(row=row_idx, column=c).border = thin_border
                if row_idx % 2 == 1:
                    ws_act.cell(row=row_idx, column=c).fill = bg_alt

        # Sheet 3: Delay Risks
        ws_risks = wb.create_sheet(title="Delay Risk Analysis")
        ws_risks.views.sheetView[0].showGridLines = True
        risk_headers = ["Activity Code", "Activity Name", "Predicted Risk Score", "Risk Band", "Forecast Delay (Days)", "Top Factors"]
        for col_idx, rh in enumerate(risk_headers, start=1):
            cell = ws_risks.cell(row=1, column=col_idx, value=rh)
            cell.fill = header_fill
            cell.font = header_font

        risks = data.get("risks", [])
        for row_idx, r in enumerate(risks, start=2):
            ws_risks.cell(row=row_idx, column=1, value=str(r.get("code") or ""))
            ws_risks.cell(row=row_idx, column=2, value=str(r.get("name") or ""))
            ws_risks.cell(row=row_idx, column=3, value=float(r.get("risk_score", 0.0)))
            ws_risks.cell(row=row_idx, column=4, value=str(r.get("risk_band") or "LOW"))
            ws_risks.cell(row=row_idx, column=5, value=int(r.get("forecast_delay_days", 0)))
            ws_risks.cell(row=row_idx, column=6, value=str(r.get("top_factors") or ""))
            for c in range(1, 7):
                ws_risks.cell(row=row_idx, column=c).border = thin_border

        # Adjust column widths automatically
        for sheet in [ws_summary, ws_act, ws_risks]:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
